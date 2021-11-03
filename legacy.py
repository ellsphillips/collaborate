# Importing packages for use
import pandas as pd
import random
import openpyxl
from openpyxl import Workbook
#These will need to be changed to where you want to save and import the data.
#Path import folder should have 2 excel files. One is the list of memebers whilst the other is the list of previous matches 
Path_import = r'C:\Users\mcferr\OneDrive - Office for National Statistics\Docs\Monthly_Matches_New'
Path_export = r'C:\Users\mcferr\OneDrive - Office for National Statistics\Docs\Monthly_Matches_New'


#%%
# Importing the latest Collaborate response information. Please make sure you have downloaded the 
# latest one and that there are an even number of matches and no duplication, 
# odd numbers will break this piece of code. Please also validate Team inputs as some people write the 
#same team names in different ways.

#This is the list of memembers 
collaborate_data = pd.read_excel(Path_import+ r'\1_Collaborate_Signup.xlsx')



#%%
#This is Theo's code it aims to combine and simplfy the matching and checking previous matched stages of the code. 

#This the previous matches sheet. With the first run you still need a create a blank excel spreadsheet for it to work. 
previous_matches=pd.read_excel(Path_import+r'\2_Collaborate_previous_matches.xlsx',index_col=0)

#We make a unquie indentifier for each match, this is just a merging of the emails of the partcipants this is the alphbetical ordering of their emails
previous_matches['check_string'] = previous_matches.apply(lambda row: ''.join(sorted([row['Email Address 1'], row['Email Address 2']])), axis=1)
#We make a list of emails
#Change 'Email address (please use your ONS account)' to whatever the Email column is called in your sign up form
emails=list(collaborate_data['Email address (please use your ONS account)'])
#This bit of code will check for duplicates and print them out.
#It will then remove these duplicates but it is good to check in the sign up sheet that they are in fact duplicates
if len(emails) != len(set(emails)):
    print(set([x for x in emails if emails.count(x) > 1]))
    emails=list(set(emails))

#initating an empty list for the matches to go into
matches=[]
#We first need to make an even number of emails. This will remove one email if there is an odd number
if len(emails)%2==1:
          emails.remove('richard.mcferran@ons.gov.uk')
#We select at random an email, we then randomly select an email to match with it. 
#If this pair has occured previously we reselect until we get a new unquie match          
#Because we replace the "failed" match back into the set there is a potential to reselect this match again and again.
#So technically the proccess could go on forever. But in praticallity it terminates in a fraction of a second.
while len(emails)>0:
    index1=random.randint(0,len(emails)-1)
    email1=emails[index1]
    emails.pop(index1)
    unique=False
    while unique==False:
        index2=random.randint(0,len(emails)-1)
        email2=emails[index2]
        email_pair=[email1,email2]
        UI=''.join(sorted(email_pair))
        unique=not any(n in UI for n in previous_matches.iloc[:,2])
    emails.pop(index2)
    matches.append(email_pair)
    

matches1=pd.DataFrame(matches, columns={'Email Address 1','Email Address 2'})
matches2=matches1.copy()
matches2.rename(columns={'Email Address 2':'Email Address 1','Email Address 1':'Email Address 2'},inplace=True);
    #we now merge the 2 data frames, this creates a version with both orderings of the matches
frames=[matches2,matches1]
matches=pd.concat(frames).copy()
#make all emails lower case
matches['Email Address 1']=matches['Email Address 1'].str.lower()
    #We now make it alphebetical order
matches.sort_values(by=['Email Address 1'],inplace=True)

#%%
# We run old code to double check that this works
# This bit of code is preparing to see whether matches have occurred in previous months
# by checking it against the master collaborate matches spreadsheet


matches['check_string'] = matches.apply(lambda row: ''.join(sorted([row['Email Address 1'], row['Email Address 2']])), axis=1)
already_matched = matches.merge(previous_matches, on='check_string', how='left')
already_matched = already_matched.dropna()
#its empty which is good

#%% New way of outputing
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
#outputing the new matches
Matches_table=load_workbook(Path_export+r'\5_Collaborate_Matches_Table.xlsx')
del Matches_table['Python Output New']
Output_sheet=Matches_table.create_sheet('Python Output New')
Output_sheet.title='Python Output New'
for r in dataframe_to_rows(matches, index=True, header=True):
    Output_sheet.append(r)
#We also output the new collaborate list 
del Matches_table['Collaborate list']
Output_sheet=Matches_table.create_sheet('Collaborate list')
Output_sheet.title='Collaborate list'
for r in dataframe_to_rows(collaborate_data, index=True, header=True):
    Output_sheet.append(r)
Matches_table.save(filename=Path_export+r'\5_Collaborate_Matches_Table.xlsx')   

#Add the new matches onto previous_matches
previous_matches=pd.read_excel(Path_import+r'\2_Collaborate_previous_matches.xlsx')
previous_matches=pd.concat([previous_matches,matches1])
writer = pd.ExcelWriter(Path_export+r'\2_Collaborate_previous_matches.xlsx', engine='xlsxwriter')
previous_matches.to_excel(writer)
writer.save()
writer.close()
#%%
# This code is redundent
#emails = list(collaborate_data['Email address (please use your ONS account)'])

# num_groups = int(len(emails)/2)
#We define this as a function, this is so it is easy to re run matchings if we have repeat matches is this a good way of doing it? Probably not seems wasteful. Could try 
#def random_pairs():
    #random.shuffle(emails)
    #We check to see if we have an even or odd number of respondents. If we have an odd number we can remove one of the collaborate team 
    #For now I have just removed me but we should have some sort of rotation system for this I think
   # if len(emails)%2==1:
          #  emails.remove('Theo.Hlustik-Smith@ons.gov.uk')
  # matches = [emails[x:x+2] for x in range(0, len(emails), 2)]
    #matches1= pd.DataFrame(matches, columns={'Match Two','Match One'})
   # matches2=matches1.copy();
   # matches2.rename(columns={'Match Two':'Match One','Match One':'Match Two'},inplace=True);
    #we now merge the 2 data frames, this creates a version with both orderings of the matches
   # frames=[matches2,matches1]
   # matches=pd.concat(frames)
    #We now make it alphebetical order
    #matches=matches.sort_values('Match One');
#%%
#OLD OUTPUT METHOD 
#Add the new matches onto previous_matches
#previous_matches=pd.read_excel(Path_import+r'\2_Collaborate_previous_matches.xlsx')
#previous_matches=pd.concat([previous_matches,matches1])
#writer = pd.ExcelWriter(Path_export+r'\2_Collaborate_previous_matches.xlsx', engine='xlsxwriter')
#previous_matches.to_excel(writer)
#writer.save()
#writer.close()

#writer = pd.ExcelWriter(Path_export+r'\4_Python_Output.xlsx', engine='xlsxwriter')
#matches.to_excel(writer)
#writer.save()
#writer.close()
